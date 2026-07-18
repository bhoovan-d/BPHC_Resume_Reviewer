"""
Resume Reviewer - Cerebras Version
===================================
Resumes evaluation of candidates skipping any that already have a score.
Uses local PDF text extraction (pypdf) + Cerebras LLM for evaluation.
No file upload needed -> much faster than the Gemini approach.

Usage:
    python review_with_cerebras.py --output Reviewed_50_Candidates.xlsx [--limit 50]
"""

import os, sys, json, time, requests, re, io, argparse
import pypdf

if sys.platform.startswith('win'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from cerebras.cloud.sdk import Cerebras
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# -- Config -------------------------------------------------------------------
EXCEL_FILE     = r"Enrolled_Students_List_For_Placement_HYD_Summer_Internship_2026-27_Resumes.xlsx"
OUTPUT_FILE    = r"Reviewed_50_Candidates.xlsx"
CEREBRAS_MODEL = "gpt-oss-120b"

DATA_START_ROW = 4
COL_SNO     = 1
COL_NAME    = 2
COL_ROLL    = 3
COL_LINK    = 4
COL_JPT     = 5
COL_REMARKS = 6
COL_SCORE   = 7
COL_VERDICT = 8

GREEN_FILL  = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
GREY_FILL   = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")
SCORE_THRESHOLDS = {"green": 80, "yellow": 55}

# -- System prompt ------------------------------------------------------------
SYSTEM_PROMPT = """You are an expert resume reviewer for engineering student placements (branches include electronics, IT, and general engineering).
Resumes are submitted through Superset, a placement platform, so formatting/layout is enforced by the platform. Focus your evaluation on content, not layout. However, the resume must be strictly one page.

You will receive:
1. A rubric explaining the evaluation criteria.
2. The extracted text content of a student's resume.
3. Page count of the resume PDF (must be 1 page).

Your task:
- Carefully evaluate the resume content based on the 100-point rubric below.
- Note: The current evaluation month and year are provided. Do not flag dates as "future dates" unless strictly later than the current evaluation date.
- Output a JSON object with exactly these keys:
  {
    "score": <integer 0-100>,
    "verdict": <"green" | "yellow" | "red">,
    "remarks": "<strict remarks following the Remark Writing Rules below>"
  }

RUBRIC (100 pts total):
1. Structure/Field Completeness (10 pts)
   - All relevant Superset fields filled meaningfully, no N/A or blank sections.
2. Content Quality -- Action Verbs & Quantification (30 pts)
   - Bullets start with strong action verbs, not passive phrases (10 pts)
   - Achievements are quantified with %, time, scale, etc. (10 pts)
   - Implicit STAR logic: context + action + measurable result (10 pts)
3. Projects & Experience Depth (25 pts)
   - Real technical/functional depth, not surface-level tech listing (10 pts)
   - Relevance to target roles (10 pts)
   - Internships/PORs show progression or ownership (5 pts)
4. Skills & Academic Fit (10 pts)
   - Skills are relevant and specific, not generic keyword-stuffing (5 pts)
   - CGPA/coursework represented appropriately (5 pts)
5. Overall Human Read -- Gut Check (25 pts)
   - Feels polished, confident, ready to send as-is (12 pts)
   - Tells a coherent story, not a mechanical checklist (13 pts)

RED FLAGS -- deduct 5 pts EACH, stacks, no cap:
   - Unprofessional email address
   - Unexplained timeline gaps
   - Factual inconsistencies (dates in the future)
   - A core section left essentially empty
   - Obvious leftover placeholder/template text
   - Content clearly overcrowded/cramped to force-fit the one-page limit
   - Resume is more than 1 page (deduct 5 pts per extra page)

COLOR MAPPING:
   - green  : 80-100  (Excellent/Good)
   - yellow : 55-79   (Average/Needs improvement)
   - red    : below 55 (Poor)

REMARK WRITING RULES (strict):
   - Length: 2-3 sentences, under ~50 words total.
   - Every sentence must reference something SPECIFIC from this resume (a project name, a skill, a missing section).
   - Structure:
     (1) Sentence 1: what is working, if anything is genuinely strong
     (2) Sentence 2: the single biggest gap, naming which rubric category
     (3) Sentence 3: one concrete, actionable fix
   - Do NOT restate the score or color inside the remark.
   - If red flags were triggered, name which ones explicitly.

Output ONLY the JSON object, no extra text."""


# -- Helpers ------------------------------------------------------------------
def download_pdf(url, timeout=30):
    try:
        resp = requests.get(url, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
        return resp.content
    except Exception as e:
        print(f"   Warning: Download failed: {e}")
        return None


def extract_text_from_pdf(pdf_bytes):
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        page_count = len(reader.pages)
        text = "\n\n".join(page.extract_text() or "" for page in reader.pages).strip()
        return text, page_count
    except Exception as e:
        return f"[Could not extract text: {e}]", 1


def clean_json_text(text):
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
    text = text.strip()
    if text.startswith('{'):
        depth = 0
        for idx, char in enumerate(text):
            if char == '{': depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    return text[:idx+1]
    return text


def evaluate_resume_cerebras(client, resume_text, page_count, student_name):
    from datetime import datetime
    current_date = datetime.now().strftime("%B %Y")
    user_prompt = f"""Please evaluate the resume of student: {student_name}.
Current date context: {current_date}. All dates up to {current_date} are past/present, NOT future.
Page count: {page_count} page(s).

--- RESUME TEXT START ---
{resume_text}
--- RESUME TEXT END ---

Return ONLY a valid JSON object with keys: score, verdict, remarks."""

    try:
        response = client.chat.completions.create(
            model=CEREBRAS_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_prompt},
            ],
            max_tokens=500,
            temperature=0.2,
        )
        text = response.choices[0].message.content.strip()
        cleaned = clean_json_text(text)
        result = json.loads(cleaned)
        score   = int(result.get("score", 0))
        verdict = str(result.get("verdict", "red")).lower()
        remarks = str(result.get("remarks", "No remarks generated."))
        if score >= SCORE_THRESHOLDS["green"]:   verdict = "green"
        elif score >= SCORE_THRESHOLDS["yellow"]: verdict = "yellow"
        else:                                     verdict = "red"
        return {"score": score, "verdict": verdict, "remarks": remarks}
    except json.JSONDecodeError as e:
        return {"score": 0, "verdict": "red", "remarks": f"[JSON parse error: {e}]"}
    except Exception as e:
        return {"score": 0, "verdict": "red", "remarks": f"[Evaluation error: {e}]"}


def get_fill(verdict):
    return {"green": GREEN_FILL, "yellow": YELLOW_FILL, "red": RED_FILL}.get(verdict, GREY_FILL)


def apply_results_to_sheet(ws, row_idx, result):
    fill = get_fill(result["verdict"])
    ws.cell(row=row_idx, column=COL_NAME).fill = fill
    ws.cell(row=row_idx, column=COL_REMARKS).value = result["remarks"]
    ws.cell(row=row_idx, column=COL_SCORE).value   = result["score"]
    verdict_cell = ws.cell(row=row_idx, column=COL_VERDICT)
    verdict_cell.value = result["verdict"].capitalize()
    verdict_cell.fill  = fill


def safe_save(wb, filename):
    try:
        wb.save(filename)
    except PermissionError:
        base, ext = os.path.splitext(filename)
        for i in range(1, 100):
            alt = f"{base}_{i}{ext}"
            try:
                wb.save(alt)
                print(f"   Saved backup to: {alt}")
                return
            except PermissionError:
                continue
        print("   Failed to save file!")


def add_headers(ws):
    ws.cell(row=3, column=COL_REMARKS).value = "AI Remarks"
    ws.cell(row=3, column=COL_SCORE).value   = "Score"
    ws.cell(row=3, column=COL_VERDICT).value = "Verdict"
    for col in [COL_REMARKS, COL_SCORE, COL_VERDICT]:
        ws.cell(row=3, column=col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(COL_REMARKS)].width = 80
    ws.column_dimensions[get_column_letter(COL_SCORE)].width   = 10
    ws.column_dimensions[get_column_letter(COL_VERDICT)].width  = 12


# -- Main ---------------------------------------------------------------------
def main():
    print("=" * 60)
    print("  Resume Reviewer (Cerebras) -- Fast Evaluation Tool")
    print("=" * 60)

    parser = argparse.ArgumentParser()
    parser.add_argument("--limit",  type=int, default=50)
    parser.add_argument("--output", type=str, default=OUTPUT_FILE)
    args, _ = parser.parse_known_args()
    output_file = args.output

    api_key = os.getenv("CEREBRAS_API_KEY", "").strip()
    if not api_key:
        api_key = input("Cerebras API key: ").strip()
    if not api_key:
        print("No API key. Exiting.")
        sys.exit(1)

    client = Cerebras(api_key=api_key)
    print(f"Model: {CEREBRAS_MODEL}")

    # Load existing output if it exists (to resume), else load source
    if os.path.exists(output_file):
        print(f"\nLoading existing output: {output_file} (will skip already-evaluated rows)")
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
    else:
        print(f"\nCreating new output from source: {EXCEL_FILE}")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

    add_headers(ws)

    # Collect all candidate rows up to limit
    all_rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        name_cell = row[COL_NAME - 1]
        link_cell = row[COL_LINK - 1]
        name = name_cell.value
        if not name or str(name).strip() == "":
            continue
        url = link_cell.hyperlink.target if link_cell.hyperlink else None
        all_rows.append((row[0].row, str(name).strip(), url))
    all_rows = all_rows[:args.limit]

    # Skip already-evaluated rows
    to_process = []
    already_done = 0
    for (row_num, name, url) in all_rows:
        existing_score = ws.cell(row=row_num, column=COL_SCORE).value
        if existing_score is not None:
            already_done += 1
        else:
            to_process.append((row_num, name, url))

    total = len(to_process)
    print(f"\nAlready evaluated: {already_done}")
    print(f"Remaining to evaluate: {total}")
    print("-" * 60)

    if total == 0:
        print("All candidates already evaluated!")
        safe_save(wb, output_file)
        return

    processed = skipped = errors = 0

    for idx, (row_num, student_name, url) in enumerate(to_process, 1):
        print(f"[{idx}/{total}] {student_name}")

        if not url:
            print("   No resume link -- skipping.")
            ws.cell(row=row_num, column=COL_NAME).fill = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "No resume link available."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Skipped"
            skipped += 1
            safe_save(wb, output_file)
            continue

        print(f"   Downloading PDF...")
        pdf_bytes = download_pdf(url)
        if not pdf_bytes:
            print("   Download failed -- skipping.")
            ws.cell(row=row_num, column=COL_NAME).fill = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "Resume download failed."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Error"
            errors += 1
            safe_save(wb, output_file)
            continue

        print(f"   Extracting text...")
        resume_text, page_count = extract_text_from_pdf(pdf_bytes)
        print(f"   Pages: {page_count}, Text: {len(resume_text)} chars")

        print(f"   Evaluating with Cerebras ({CEREBRAS_MODEL})...")
        result = evaluate_resume_cerebras(client, resume_text, page_count, student_name)
        print(f"   Score: {result['score']} | Verdict: {result['verdict'].upper()}")
        print(f"   {result['remarks'][:100]}...")

        apply_results_to_sheet(ws, row_num, result)
        processed += 1
        safe_save(wb, output_file)
        print()

    safe_save(wb, output_file)

    print("=" * 60)
    print(f"DONE! Results saved to: {output_file}")
    print(f"   Processed : {processed}")
    print(f"   Skipped   : {skipped}")
    print(f"   Errors    : {errors}")
    print("=" * 60)


if __name__ == "__main__":
    main()
