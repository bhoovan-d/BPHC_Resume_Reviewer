"""
Resume Reviewer Automation
==========================
Reads an Excel sheet of students + resume links, reviews each resume
against a rubric PDF using Google Gemini (via new google-genai SDK),
and writes color-coded results + remarks back to the Excel file.

Usage:
    python review_resumes.py

    You will be prompted to enter any personal/additional rules
    interactively before processing begins.
"""

import os
import sys
import json
import time
import tempfile
import requests
import re
import io
if sys.platform.startswith('win'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import PatternFill, Font, colors
from openpyxl.styles import Font as XLFont
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ── Config ──────────────────────────────────────────────────────────────────
load_dotenv()

EXCEL_FILE     = r"bhoovan_reviews.xlsx"
RUBRIC_PDF     = r"Resume Review Doc.pdf"
OUTPUT_FILE    = r"bhoovan_reviews_output.xlsx"  # default; can be overridden via --output arg

DATA_START_ROW = 2          # Row where student data begins (1-indexed) after inserting headers
COL_NAME       = 1          # A
COL_ROLL       = 2          # B
COL_LINK       = 3          # C  (hyperlink)
COL_JPT        = 4          # D
COL_REMARKS    = 5          # E  — will be overwritten with AI remarks
COL_SCORE      = 6          # F  — new column for numeric score
COL_VERDICT    = 7          # G  — new column for Green/Yellow/Red label

# Color fills
GREEN_FILL  = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
GREY_FILL   = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")

SCORE_THRESHOLDS = {"green": 90, "yellow": 55}   # ≥90 green, ≥55 yellow, else red

# ── Gemini setup ─────────────────────────────────────────────────────────────
def setup_gemini(api_key: str):
    return genai.Client(
        api_key=api_key,
        http_options=types.HttpOptions(timeout=60_000)
    )


def upload_rubric(client, rubric_path: str):
    """Upload rubric PDF to Gemini Files API (once per run)."""
    print(f"📄 Uploading rubric PDF to Gemini...")
    rubric_file = client.files.upload(file=rubric_path)
    print(f"   ✅ Rubric uploaded: {rubric_file.name}")
    return rubric_file


# ── Resume downloading ────────────────────────────────────────────────────────
def download_pdf(url: str, timeout: int = 30) -> bytes | None:
    """Download a PDF from a URL. Returns raw bytes or None on failure."""
    try:
        resp = requests.get(url, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
        return resp.content
    except Exception as e:
        print(f"   ⚠️  Download failed: {e}")
        return None


# ── AI Evaluation ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert resume reviewer for engineering student placements (branches include electronics, IT, and general engineering — apply judgment broadly across fields).
Resumes are submitted through Superset, a placement platform — so formatting/layout is largely enforced by the platform, not the student. Focus your evaluation on content, not layout. However, the resume must be strictly one page.

You will receive:
1. A rubric PDF explaining the evaluation criteria (formatting, BITS template rules, keywords, action verb guidance).
2. A student's resume PDF.
3. (Optionally) Additional personal rules from the reviewer.

Your task:
- Carefully evaluate the resume content based on the 100-point rubric below, deducting for any red flags, and determine the score and verdict.
- TIMELINE RULE: Students graduating in 2027–2028 will naturally have internships and projects dated 2024–2026. Do NOT flag these as gaps or inconsistencies. Only flag a date as a red flag if it is strictly AFTER the current evaluation month/year provided in the user request.
- Output a JSON object with exactly these keys:
  {{
    "score": <integer 0-100>,
    "verdict": <"green" | "yellow" | "red">,
    "remarks": "<strict remarks following the Remark Writing Rules below>"
  }}

RUBRIC (100 pts total):
1. Structure/Field Completeness (10 pts)
   - All relevant Superset fields filled meaningfully — no "N/A," blank sections, or leftover placeholder text where real content should be.
2. Content Quality — Action Verbs & Quantification (30 pts)
   - Bullets start with strong action verbs, not passive phrases like "responsible for / helped / worked on" (10 pts)
   - Achievements are quantified — %, ₹/$, time, scale — not vague claims (10 pts)
   - Implicit STAR logic present: context + action + measurable result, even if not rigidly labeled (10 pts)
3. Projects & Experience Depth (25 pts)
   - Real technical/functional depth, not a surface-level listing of technologies (10 pts)
   - Relevance to target roles (electronics/IT-adjacent or general engineering) (10 pts)
   - Internships/PORs show progression or ownership, not just attendance (5 pts)
4. Skills & Academic Fit (10 pts)
   - Skills are relevant and specific, not generic keyword-stuffing (5 pts)
   - CGPA/coursework represented appropriately (5 pts)
5. Overall Human Read — Gut Check (25 pts)
   Step back and ask: "If I were a recruiter skimming this for 20 seconds, would it land well?"
   - Feels polished, confident, ready to send as-is (12 pts)
   - Tells a coherent story about the candidate, rather than reading like a checklist was filled in mechanically (13 pts)
   *Note: This section is deliberately holistic. If the mechanical score from sections 1-4 doesn't match how the resume actually reads overall, let this section pull the total up or down by a few points to correct for it. A resume that's slightly imperfect on individual bullets but reads strong, clean, and intentional as a whole should score WELL here — do not penalize the same minor flaw twice.

RED FLAGS — deduct 5 pts EACH, stacks, no cap:
   - Unprofessional email address
   - Dates strictly in the future (after the current evaluation month/year) — do NOT flag past internships or projects as future dates just because they occurred before the student graduates
   - A core section left essentially empty
   - Obvious leftover placeholder/template text
   - Content clearly overcrowded/cramped to force-fit the one-page limit

COLOR MAPPING & VERDICTS (based on final total after deductions):
   - green  : 90-100  (Excellent/Good)
   - yellow : 55-89   (Average/Needs improvement)
   - red    : below 55 (Poor)

REMARK WRITING RULES (strict):
   - Length: 2-3 sentences, under ~60 words total.
   - Remarks must focus on COSMETIC and ACTIONABLE improvements the student can make — NOT on judging the content or relevance (scoring handles that). Prioritise these categories in order:
     * Bullet formatting: are bullets starting with strong action verbs? Name the specific weak verb and suggest a stronger one.
     * Quantification: are achievements backed by numbers (%, time saved, scale)? Name the specific bullet missing a metric.
     * STAR format: is there context + action + result? Name the specific project/internship where the result is missing.
     * Other cosmetic fixes: missing sections, placeholder text left in, crowded layout, unprofessional email.
   - Every sentence must reference something SPECIFIC from this resume's actual text (project name, bullet text, section name) — never write generic advice like "add more metrics" without naming where.
   - Structure:
     (1) Sentence 1: name the single strongest cosmetic positive (e.g. "Bullets in the XYZ project effectively start with action verbs and include metrics.")
     (2) Sentence 2: name the single most impactful cosmetic fix needed, citing the exact section/bullet.
     (3) Sentence 3: give the concrete rewrite or fix (e.g. "Replace 'Worked on API' with 'Built REST API handling 10k req/s, reducing latency by 30%'.")
   - Do NOT restate the score or color inside the remark.
   - Do NOT comment on whether the content/domain is relevant or impressive — only on how it is written and formatted.
   - If red flags were triggered, name which one(s) explicitly in Sentence 2 instead.

Output ONLY the JSON object, no extra text.
"""


def call_with_retry(func, *args, max_retries=5, initial_backoff=2, **kwargs):
    backoff = initial_backoff
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            err_msg = str(e).lower()
            is_temporary = any(code in err_msg for code in ["503", "429", "unavailable", "rate_limit", "resource_exhausted", "quota", "overloaded"])
            if is_temporary and attempt < max_retries - 1:
                print(f"   ⚠️  Temporary API error (attempt {attempt+1}/{max_retries}): {e}. Retrying in {backoff} seconds...")
                time.sleep(backoff)
                backoff *= 2
            else:
                raise e

def clean_json_text(text):
    text = text.strip()
    # Remove markdown code fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
    text = text.strip()
    
    # Extract only the first valid JSON block
    if text.startswith('{'):
        depth = 0
        for idx, char in enumerate(text):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    return text[:idx+1]
    return text

def evaluate_resume(
    client,
    rubric_file,
    resume_bytes: bytes,
    student_name: str,
    personal_rules: str = "",
) -> dict:
    """Send resume to Gemini for evaluation. Returns dict with score/verdict/remarks."""

    # Write resume bytes to a temp file and upload
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(resume_bytes)
        tmp_path = tmp.name

    try:
        resume_file = call_with_retry(client.files.upload, file=tmp_path)

        from datetime import datetime
        current_date = datetime.now().strftime("%B %Y")
        extra = f"\n\nAdditional reviewer rules:\n{personal_rules}" if personal_rules.strip() else ""

        prompt = f"""Please evaluate the resume of student: {student_name}.
Current date context for evaluation: {current_date}. All dates on the resume up to this date (including {current_date}) are in the past or present, NOT the future.

Use the rubric PDF provided and the scoring guide in your instructions.{extra}

Return ONLY a JSON object."""

        response = call_with_retry(
            client.models.generate_content,
            model="gemini-3.1-flash-lite",
            contents=[rubric_file, resume_file, prompt],
            config=types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
                response_mime_type="application/json"
            )
        )
        text = response.text.strip()
        cleaned_text = clean_json_text(text)
        result = json.loads(cleaned_text)

        # Validate and normalise
        score   = int(result.get("score", 0))
        verdict = str(result.get("verdict", "red")).lower()
        remarks = str(result.get("remarks", "No remarks generated."))

        # Ensure verdict matches score
        if score >= SCORE_THRESHOLDS["green"]:
            verdict = "green"
        elif score >= SCORE_THRESHOLDS["yellow"]:
            verdict = "yellow"
        else:
            verdict = "red"

        return {"score": score, "verdict": verdict, "remarks": remarks}

    except json.JSONDecodeError as e:
        print(f"   ⚠️  JSON parse error: {e}. Raw: {text[:200]}")
        return {"score": 0, "verdict": "red", "remarks": f"[Evaluation error: could not parse AI response] {text[:200]}"}
    except Exception as e:
        print(f"   ⚠️  Gemini error: {e}")
        return {"score": 0, "verdict": "red", "remarks": f"[Evaluation error: {e}]"}
    finally:
        os.unlink(tmp_path)


# ── Excel helpers ──────────────────────────────────────────────────────────────
def get_fill(verdict: str) -> PatternFill:
    return {"green": GREEN_FILL, "yellow": YELLOW_FILL, "red": RED_FILL}.get(verdict, GREY_FILL)


def apply_results_to_sheet(ws, row_idx: int, result: dict):
    """Color the name cell and write score/verdict/remarks."""
    fill = get_fill(result["verdict"])

    # Color the Name cell (col B)
    name_cell = ws.cell(row=row_idx, column=COL_NAME)
    name_cell.fill = fill

    # Write Remarks (col F)
    ws.cell(row=row_idx, column=COL_REMARKS).value = result["remarks"]

    # Write Score (col G)
    ws.cell(row=row_idx, column=COL_SCORE).value = result["score"]

    # Write Verdict label (col H)
    verdict_cell = ws.cell(row=row_idx, column=COL_VERDICT)
    verdict_cell.value = result["verdict"].capitalize()
    verdict_cell.fill = fill


def add_headers(ws):
    """Ensure headers exist in row 1."""
    ws.cell(row=1, column=COL_NAME).value    = "Name"
    ws.cell(row=1, column=COL_ROLL).value    = "Roll No"
    ws.cell(row=1, column=COL_LINK).value    = "Primary Resume"
    ws.cell(row=1, column=COL_JPT).value     = "Reviewer"
    ws.cell(row=1, column=COL_REMARKS).value = "AI Remarks"
    ws.cell(row=1, column=COL_SCORE).value   = "Score"
    ws.cell(row=1, column=COL_VERDICT).value = "Verdict"

    # Bold headers
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)

    # Set reasonable column width for remarks
    ws.column_dimensions[get_column_letter(COL_REMARKS)].width = 80
    ws.column_dimensions[get_column_letter(COL_SCORE)].width   = 10
    ws.column_dimensions[get_column_letter(COL_VERDICT)].width  = 12


def safe_save(wb, filename):
    try:
        wb.save(filename)
    except PermissionError:
        print(f"   ⚠️  Permission denied when writing to {filename}. Is it open in Excel?")
        # Try to save with an alternative name
        base, ext = os.path.splitext(filename)
        for i in range(1, 100):
            alt_name = f"{base}_{i}{ext}"
            try:
                wb.save(alt_name)
                print(f"   ✅ Saved progress backup to: {alt_name}")
                return
            except PermissionError:
                continue
        print("   ❌ Failed to save even to backup files!")

# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  📋  Resume Reviewer — Automated Evaluation Tool")
    print("=" * 60)

    import argparse
    parser = argparse.ArgumentParser(description="Automate resume reviews.")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of students to review")
    parser.add_argument("--rules", type=str, default=None, help="One-time personal rules for the evaluation")
    parser.add_argument("--output", type=str, default=None, help="Override output Excel filename")
    
    # Use parse_known_args in case run with unexpected flags in IDE test
    args, unknown = parser.parse_known_args()

    # ── API Key ──
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        api_key = input("\n🔑 Enter your Google Gemini API key: ").strip()
    if not api_key:
        print("❌ No API key provided. Exiting.")
        sys.exit(1)

    # ── Personal rules ──
    if args.rules is not None:
        personal_rules = args.rules.strip()
        print(f"\n📝 Using personal rules from CLI arguments ({len(personal_rules)} chars).")
    else:
        print("\n📝 Enter any additional personal rules for evaluation.")
        print("   (Press Enter twice when done, or just Enter once to skip)")
        lines = []
        while True:
            line = input("   > ")
            if line == "" and (not lines or lines[-1] == ""):
                break
            lines.append(line)
        personal_rules = "\n".join(lines).strip()
        if personal_rules:
            print(f"\n   ✅ Personal rules recorded ({len(personal_rules)} chars).")
        else:
            print("\n   ℹ️  No personal rules provided — using rubric only.")

    limit = args.limit

    # ── Output file override ──
    global OUTPUT_FILE
    if args.output:
        OUTPUT_FILE = args.output
        print(f"\n📁 Output file set to: {OUTPUT_FILE}")

    # ── Load Excel ──
    print(f"\n📊 Loading input Excel: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        sys.exit(1)
    wb_in = openpyxl.load_workbook(EXCEL_FILE)
    ws_in = wb_in.active
    print(f"   ✅ Loaded {ws_in.max_row} rows × {ws_in.max_column} cols from sheet '{ws_in.title}'")

    # Read students first from the original file (WITHOUT modifying in memory)
    all_rows = []
    for r in range(1, ws_in.max_row + 1):
        name = ws_in.cell(row=r, column=1).value
        roll = ws_in.cell(row=r, column=2).value
        link_cell = ws_in.cell(row=r, column=3)
        reviewer = ws_in.cell(row=r, column=4).value
        url = link_cell.hyperlink.target if link_cell.hyperlink else None
        
        if not name or str(name).strip() == "":
            continue
            
        all_rows.append({
            "name": str(name).strip(),
            "roll": str(roll).strip() if roll else "",
            "url": url,
            "reviewer": str(reviewer).strip() if reviewer else ""
        })
    print(f"   ✅ Successfully read {len(all_rows)} student rows.")

    # ── Setup Gemini ──
    client = setup_gemini(api_key)

    # ── Upload rubric ──
    if not os.path.exists(RUBRIC_PDF):
        print(f"❌ Rubric PDF not found: {RUBRIC_PDF}")
        sys.exit(1)
    rubric_file = upload_rubric(client, RUBRIC_PDF)

    # ── Initialize Output Workbook ──
    if os.path.exists(OUTPUT_FILE):
        print(f"📊 Loading existing output Excel: {OUTPUT_FILE}")
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        print(f"📊 Creating new output Excel: {OUTPUT_FILE}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Populate initial student data rows
        HYPERLINK_FONT = XLFont(color="0563C1", underline="single")
        for idx, student in enumerate(all_rows, 2):
            ws.cell(row=idx, column=COL_NAME).value = student["name"]
            ws.cell(row=idx, column=COL_ROLL).value = student["roll"]
            link_cell = ws.cell(row=idx, column=COL_LINK)
            link_cell.value = "Link"
            if student["url"]:
                link_cell.hyperlink = student["url"]
                link_cell.font = HYPERLINK_FONT
            ws.cell(row=idx, column=COL_JPT).value = student["reviewer"]
            
        add_headers(ws)
        wb.save(OUTPUT_FILE)

    # ── Process students ──
    total = len(all_rows)
    processed = 0
    skipped   = 0
    errors    = 0

    if limit is not None:
        all_rows = all_rows[:limit]
        total = len(all_rows)
        print(f"\n🎓 Found {total} students to review (limited to first {limit}).\n")
    else:
        print(f"\n🎓 Found {total} students to review.\n")
    print("-" * 60)

    for idx, student in enumerate(all_rows, 1):
        row_num = idx + 1 # Row index in ws (output sheet) is idx + 1
        student_name = student["name"]
        url = student["url"]
        
        # Check if already evaluated (score is not None or verdict is set)
        existing_score = ws.cell(row=row_num, column=COL_SCORE).value
        existing_verdict = ws.cell(row=row_num, column=COL_VERDICT).value
        if existing_score is not None or existing_verdict in ["Green", "Yellow", "Red"]:
            print(f"[{idx}/{total}] {student_name} — already evaluated. Skipping.")
            continue

        print(f"[{idx}/{total}] {student_name}")

        if not url:
            print("   ⚠️  No resume link found — skipping.")
            ws.cell(row=row_num, column=COL_NAME).fill = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "No resume link available."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Skipped"
            skipped += 1
            safe_save(wb, OUTPUT_FILE)
            continue

        # Download PDF
        print(f"   ⬇️  Downloading resume...")
        pdf_bytes = download_pdf(url)
        if not pdf_bytes:
            print("   ❌ Could not download resume — skipping.")
            ws.cell(row=row_num, column=COL_NAME).fill = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "Resume download failed."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Error"
            errors += 1
            safe_save(wb, OUTPUT_FILE)
            continue

        # Evaluate
        print(f"   🤖 Evaluating with Gemini...")
        result = evaluate_resume(client, rubric_file, pdf_bytes, student_name, personal_rules)
        print(f"   ✅ Score: {result['score']} | Verdict: {result['verdict'].upper()}")
        print(f"   💬 {result['remarks'][:100]}...")

        # Write to Excel
        apply_results_to_sheet(ws, row_num, result)
        processed += 1

        # Save after every student (so progress isn't lost on crash)
        safe_save(wb, OUTPUT_FILE)

        # Respect Gemini free-tier rate limit (15 req/min)
        if idx < total:
            time.sleep(6)

        print()

    # ── Final save ──
    safe_save(wb, OUTPUT_FILE)

    print("=" * 60)
    print(f"✅ DONE! Results saved to: {OUTPUT_FILE}")
    print(f"   Processed : {processed}")
    print(f"   Skipped   : {skipped}  (no link)")
    print(f"   Errors    : {errors}   (download/AI failure)")
    print("=" * 60)


if __name__ == "__main__":
    main()
