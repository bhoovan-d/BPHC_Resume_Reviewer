"""
Resume Reviewer - Fast Multi-LLM Version
=========================================
Uses Groq (primary) + NVIDIA NIM (fallback) with local PDF text extraction.
Automatically skips already-evaluated candidates and saves after each one.

Usage:
    python review_fast.py [--limit 50] [--output Reviewed_50_Candidates.xlsx]
"""

import os, sys, json, requests, re, io, argparse, time
import pypdf

if sys.platform.startswith('win'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_SOURCE = r"Enrolled_Students_List_For_Placement_HYD_Summer_Internship_2026-27_Resumes.xlsx"
OUTPUT_FILE  = r"Reviewed_50_Candidates.xlsx"

DATA_START_ROW = 4
COL_NAME    = 2
COL_LINK    = 4
COL_REMARKS = 6
COL_SCORE   = 7
COL_VERDICT = 8

GREEN_FILL  = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
GREY_FILL   = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)

# ---------------------------------------------------------------------------
# System prompt (same rubric as before)
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """You are an expert resume reviewer for engineering student placements (branches include electronics, IT, and general engineering).
Resumes are submitted through Superset, a placement platform, so formatting/layout is enforced by the platform. Focus your evaluation on content, not layout. However, the resume must be strictly one page.

You will receive the extracted text of a student's resume and its page count.

Your task: evaluate the resume based on the rubric below and output ONLY a JSON object.

RUBRIC (100 pts total):
1. Structure/Field Completeness (10 pts) - all relevant fields filled meaningfully.
2. Content Quality - Action Verbs & Quantification (30 pts)
   - Strong action verbs, not passive phrases (10 pts)
   - Quantified achievements with %, time, scale (10 pts)
   - Implicit STAR logic: context + action + measurable result (10 pts)
3. Projects & Experience Depth (25 pts)
   - Real technical depth, not surface-level tech listing (10 pts)
   - Relevance to target roles (10 pts)
   - Internships/PORs show progression or ownership (5 pts)
4. Skills & Academic Fit (10 pts)
   - Specific, relevant skills not generic keyword-stuffing (5 pts)
   - CGPA/coursework appropriately represented (5 pts)
5. Overall Human Read - Gut Check (25 pts)
   - Feels polished, confident, ready to send as-is (12 pts)
   - Tells a coherent story, not a mechanical checklist (13 pts)

RED FLAGS - deduct 5 pts EACH (stacks, no cap):
   - Unprofessional email address
   - Unexplained timeline gaps
   - Factual inconsistencies (dates in the future relative to current date)
   - A core section left essentially empty
   - Obvious leftover placeholder/template text
   - Content clearly overcrowded/cramped to force the one-page limit
   - Resume is more than 1 page (5 pts per extra page)

COLOR MAPPING (after all deductions):
   - green  : 80-100
   - yellow : 55-79
   - red    : 0-54

REMARK RULES (strict):
   - 2-3 sentences, under ~50 words total.
   - Every sentence references something SPECIFIC from this resume (project name, missing section, skill).
   - Structure: (1) what's working, (2) biggest gap + rubric category, (3) one concrete fix.
   - Do NOT restate score or color in the remark.
   - If red flags triggered, name them explicitly.

Output ONLY valid JSON, no markdown, no extra text:
{"score": <int 0-100>, "verdict": <"green"|"yellow"|"red">, "remarks": "<string>"}"""


# ---------------------------------------------------------------------------
# LLM Clients
# ---------------------------------------------------------------------------
def make_groq_client():
    from groq import Groq
    return Groq(api_key=os.getenv("GROQ_API_KEY"), timeout=20.0)


def make_nvidia_client():
    from openai import OpenAI
    return OpenAI(
        base_url="https://integrate.api.nvidia.com/v1",
        api_key=os.getenv("NVIDIA_API_KEY"),
        timeout=20.0
    )


# Round-robin counter — alternates every call so neither API hits TPM limits
_call_counter = [0]


def call_nvidia_llama31(client, system, user):
    for attempt in range(4):
        try:
            r = client.chat.completions.create(
                model="meta/llama-3.1-70b-instruct",
                messages=[{"role": "system", "content": system},
                          {"role": "user",   "content": user}],
                max_tokens=500,
                temperature=0.2,
            )
            return r.choices[0].message.content.strip()
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate_limit" in msg.lower():
                wait = 15 * (attempt + 1)
                print(f"   NVIDIA 3.1 429 rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("NVIDIA Llama 3.1 exhausted retries")


def call_nvidia_llama33(client, system, user):
    for attempt in range(4):
        try:
            r = client.chat.completions.create(
                model="meta/llama-3.3-70b-instruct",
                messages=[{"role": "system", "content": system},
                          {"role": "user",   "content": user}],
                max_tokens=500,
                temperature=0.2,
            )
            return r.choices[0].message.content.strip()
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate_limit" in msg.lower():
                wait = 15 * (attempt + 1)
                print(f"   NVIDIA 3.3 429 rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("NVIDIA Llama 3.3 exhausted retries")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
MAX_PDF_BYTES = 20 * 1024 * 1024  # 20 MB hard cap


def download_pdf(url, timeout=(10, 60)):
    """Stream download with size cap so huge scanned PDFs don't hang forever."""
    try:
        with requests.get(url, timeout=timeout, allow_redirects=True, stream=True) as resp:
            resp.raise_for_status()
            chunks = []
            total = 0
            for chunk in resp.iter_content(chunk_size=65536):
                chunks.append(chunk)
                total += len(chunk)
                if total > MAX_PDF_BYTES:
                    print(f"   PDF too large (>{MAX_PDF_BYTES//1024//1024}MB) — skipping.")
                    return None
            data = b"".join(chunks)
            print(f"   Downloaded {total/1024:.0f} KB", end=" ")
            return data
    except Exception as e:
        print(f"   Download failed: {e}")
        return None


def extract_text(pdf_bytes):
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        pages = len(reader.pages)
        text = "\n\n".join(p.extract_text() or "" for p in reader.pages).strip()
        return text, pages
    except Exception as e:
        return f"[PDF extraction error: {e}]", 1


def clean_json(raw):
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"\s*```$", "", raw, flags=re.MULTILINE)
    raw = raw.strip()
    # Extract just the first complete JSON object
    if raw.startswith("{"):
        depth = 0
        for i, ch in enumerate(raw):
            if ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return raw[:i+1]
    return raw


def parse_result(raw_text):
    cleaned = clean_json(raw_text)
    data = json.loads(cleaned)
    score   = int(data.get("score", 0))
    remarks = str(data.get("remarks", ""))
    if score >= 80:   verdict = "green"
    elif score >= 55: verdict = "yellow"
    else:             verdict = "red"
    return {"score": score, "verdict": verdict, "remarks": remarks}


def evaluate(groq_client, nvidia_client, resume_text, page_count, name):
    from datetime import datetime
    current_date = datetime.now().strftime("%B %Y")

    user_msg = f"""Evaluate the resume of student: {name}
Current date: {current_date}. All dates up to {current_date} are past/present — do NOT flag them as future.
Page count: {page_count}

--- RESUME TEXT ---
{resume_text}
---

Return ONLY valid JSON: {{"score": int, "verdict": "green"|"yellow"|"red", "remarks": "string"}}"""

    # We use Llama 3.1 70B as primary and Llama 3.3 70B as fallback, both on NVIDIA NIM
    primary_name,   primary_fn   = ("Llama3.1", call_nvidia_llama31)
    fallback_name,  fallback_fn  = ("Llama3.3", call_nvidia_llama33)

    try:
        raw = primary_fn(nvidia_client, SYSTEM_PROMPT, user_msg)
        print(f"   [{primary_name}]", end=" ")
        return parse_result(raw)
    except Exception as e:
        print(f"   {primary_name} failed ({str(e)[:50]}), trying {fallback_name}...")

    try:
        raw = fallback_fn(nvidia_client, SYSTEM_PROMPT, user_msg)
        print(f"   [{fallback_name}]", end=" ")
        return parse_result(raw)
    except Exception as e:
        print(f"   Both APIs failed: {str(e)[:60]}")
        return {"score": 0, "verdict": "red", "remarks": f"[Both APIs failed: {str(e)[:80]}]"}


def get_fill(verdict):
    return {"green": GREEN_FILL, "yellow": YELLOW_FILL, "red": RED_FILL}.get(verdict, GREY_FILL)


def write_result(ws, row, result):
    fill = get_fill(result["verdict"])
    ws.cell(row=row, column=COL_NAME).fill    = fill
    ws.cell(row=row, column=COL_REMARKS).value = result["remarks"]
    ws.cell(row=row, column=COL_SCORE).value   = result["score"]
    vc = ws.cell(row=row, column=COL_VERDICT)
    vc.value = result["verdict"].capitalize()
    vc.fill  = fill


def safe_save(wb, path):
    try:
        wb.save(path)
    except PermissionError:
        base, ext = os.path.splitext(path)
        for i in range(1, 20):
            alt = f"{base}_{i}{ext}"
            try:
                wb.save(alt)
                print(f"   Saved to: {alt} (original locked)")
                return
            except PermissionError:
                continue
        print("   ERROR: Could not save file!")


def add_headers(ws):
    ws.cell(row=3, column=COL_REMARKS).value = "AI Remarks"
    ws.cell(row=3, column=COL_SCORE).value   = "Score"
    ws.cell(row=3, column=COL_VERDICT).value = "Verdict"
    for col in [COL_REMARKS, COL_SCORE, COL_VERDICT]:
        ws.cell(row=3, column=col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(COL_REMARKS)].width = 80
    ws.column_dimensions[get_column_letter(COL_SCORE)].width   = 10
    ws.column_dimensions[get_column_letter(COL_VERDICT)].width = 12


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit",  type=int, default=50)
    parser.add_argument("--output", type=str, default=OUTPUT_FILE)
    args, _ = parser.parse_known_args()

    print("=" * 60)
    print("  Resume Reviewer  (Groq + NVIDIA NIM)")
    print("=" * 60)

    groq_client   = make_groq_client()
    nvidia_client = make_nvidia_client()
    print("Clients ready: Groq (primary) + NVIDIA NIM (fallback)")

    # Load or create output workbook
    if os.path.exists(args.output):
        print(f"\nLoading: {args.output}")
        wb = openpyxl.load_workbook(args.output)
    else:
        print(f"\nCreating from source: {EXCEL_SOURCE}")
        wb = openpyxl.load_workbook(EXCEL_SOURCE)
    ws = wb.active
    add_headers(ws)

    # Collect candidates
    all_rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        name = row[COL_NAME - 1].value
        if not name or str(name).strip() == "":
            continue
        link_cell = row[COL_LINK - 1]
        url = link_cell.hyperlink.target if link_cell.hyperlink else None
        all_rows.append((row[0].row, str(name).strip(), url))
    all_rows = all_rows[:args.limit]

    # Split: already done vs. pending
    to_do   = []
    n_done  = 0
    for row_num, name, url in all_rows:
        score_val = ws.cell(row=row_num, column=COL_SCORE).value
        if score_val is not None:
            n_done += 1
        else:
            to_do.append((row_num, name, url))

    total = len(to_do)
    print(f"\nAlready done : {n_done}")
    print(f"To evaluate  : {total}")
    print("-" * 60)

    if total == 0:
        print("Nothing to do — all candidates already evaluated!")
        safe_save(wb, args.output)
        return

    processed = errors = skipped = 0

    for idx, (row_num, name, url) in enumerate(to_do, 1):
        print(f"\n[{idx}/{total}] {name}")

        if not url:
            print("   No resume URL — marking skipped.")
            ws.cell(row=row_num, column=COL_NAME).fill    = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "No resume link provided."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Skipped"
            ws.cell(row=row_num, column=COL_SCORE).value   = 0
            skipped += 1
            safe_save(wb, args.output)
            continue

        pdf = download_pdf(url)
        if not pdf:
            ws.cell(row=row_num, column=COL_NAME).fill    = GREY_FILL
            ws.cell(row=row_num, column=COL_REMARKS).value = "Resume PDF could not be downloaded (expired or broken link)."
            ws.cell(row=row_num, column=COL_VERDICT).value = "Error"
            ws.cell(row=row_num, column=COL_SCORE).value   = 0
            errors += 1
            safe_save(wb, args.output)
            continue

        text, pages = extract_text(pdf)
        print(f"   Pages: {pages} | Text: {len(text)} chars")

        result = evaluate(groq_client, nvidia_client, text, pages, name)
        print(f"Score: {result['score']} | Verdict: {result['verdict'].upper()}")
        print(f"   {result['remarks'][:100]}{'...' if len(result['remarks'])>100 else ''}")

        write_result(ws, row_num, result)
        processed += 1
        safe_save(wb, args.output)
        time.sleep(1)  # 1s gap to respect per-minute token limits

    safe_save(wb, args.output)

    print("\n" + "=" * 60)
    print(f"COMPLETE! Saved to: {args.output}")
    print(f"  Processed : {processed}")
    print(f"  Skipped   : {skipped}  (no URL)")
    print(f"  Errors    : {errors}   (download failed)")
    print("=" * 60)


if __name__ == "__main__":
    main()
