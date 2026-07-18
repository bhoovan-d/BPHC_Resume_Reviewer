"""
Resume Reviewer - Repair Tool
==============================
Finds rows in the Excel sheet that timed out (Score: 0, Verdict: Red, Remarks containing 'timed out')
and re-runs them using Llama 3.1 70b with a longer 45s timeout.
"""

import os, sys, json, time, requests, re, io
import pypdf
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

OUTPUT_FILE  = r"Reviewed_50_Candidates.xlsx"

COL_NAME    = 2
COL_LINK    = 4
COL_REMARKS = 6
COL_SCORE   = 7
COL_VERDICT = 8

GREEN_FILL  = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

from review_fast import SYSTEM_PROMPT, parse_result, get_fill, write_result, safe_save, download_pdf, extract_text

def call_nvidia_llama31_long(client, system, user):
    for attempt in range(3):
        try:
            r = client.chat.completions.create(
                model="meta/llama-3.1-70b-instruct",
                messages=[{"role": "system", "content": system},
                          {"role": "user",   "content": user}],
                max_tokens=500,
                temperature=0.2,
            )
            return r.choices[0].message.content.strip()
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate_limit" in msg.lower():
                wait = 20 * (attempt + 1)
                print(f"   NVIDIA 3.1 429 rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("NVIDIA Llama 3.1 long exhausted retries")

def main():
    print("=" * 60)
    print("  Resume Reviewer - Repair Tool")
    print("=" * 60)

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Find rows to repair
    to_repair = []
    for r in range(4, 54):
        name = ws.cell(r, COL_NAME).value
        if not name:
            continue
        score = ws.cell(r, COL_SCORE).value
        verdict = str(ws.cell(r, COL_VERDICT).value or "")
        remarks = str(ws.cell(r, COL_REMARKS).value or "")

        if score == 0 and "timed out" in remarks.lower():
            link_cell = ws.cell(r, COL_LINK)
            url = link_cell.hyperlink.target if link_cell.hyperlink else None
            to_repair.append((r, str(name).strip(), url))

    print(f"Found {len(to_repair)} rows that timed out:")
    for r, name, _ in to_repair:
        print(f"  Row {r}: {name}")
    print("-" * 60)

    if not to_repair:
        print("Nothing to repair!")
        return

    # Use NVIDIA client with a generous 45s timeout
    client = OpenAI(
        base_url="https://integrate.api.nvidia.com/v1",
        api_key=os.getenv("NVIDIA_API_KEY"),
        timeout=45.0
    )

    for idx, (row_num, name, url) in enumerate(to_repair, 1):
        print(f"[{idx}/{len(to_repair)}] Repairing {name}...")

        if not url:
            print("   No url — skipping.")
            continue

        pdf = download_pdf(url)
        if not pdf:
            print("   Download failed — skipping.")
            continue

        text, pages = extract_text(pdf)
        print(f"   Pages: {pages} | Text length: {len(text)}")

        from datetime import datetime
        current_date = datetime.now().strftime("%B %Y")
        user_msg = f"""Evaluate the resume of student: {name}
Current date: {current_date}. All dates up to {current_date} are past/present — do NOT flag them as future.
Page count: {pages}

--- RESUME TEXT ---
{text}
---

Return ONLY valid JSON: {{"score": int, "verdict": "green"|"yellow"|"red", "remarks": "string"}}"""

        try:
            print("   Calling Llama 3.1 with 45s timeout...")
            raw = call_nvidia_llama31_long(client, SYSTEM_PROMPT, user_msg)
            result = parse_result(raw)
            print(f"   Success! Score: {result['score']} | Verdict: {result['verdict'].upper()}")
            print(f"   Remarks: {result['remarks'][:100]}...")
            
            # Clear old error fill and write new results
            ws.cell(row=row_num, column=COL_NAME).fill = PatternFill(fill_type=None)
            write_result(ws, row_num, result)
            safe_save(wb, OUTPUT_FILE)
            print("   Saved.")
        except Exception as e:
            print(f"   FAILED to repair: {e}")

        print()

    print("Repair complete!")

if __name__ == "__main__":
    main()
