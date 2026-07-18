r"""
Quick test — evaluates ONLY the first student to validate the pipeline.
Run this before the full batch to confirm everything works.

Usage:
    C:\Users\bhoov\AppData\Local\Programs\Python\Python312\python.exe test_one.py
"""

import os, sys, json, tempfile, re, requests, io
if sys.platform.startswith('win'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from google import genai
from google.genai import types
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE = r"bhoovan_reviews.xlsx"
RUBRIC_PDF = r"Resume Review Doc.pdf"
DATA_START_ROW = 1
COL_NAME = 1
COL_LINK = 3

def clean_json_text(text):
    text = text.strip()
    # Remove markdown code fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
    text = text.strip()
    
    # Extract only the first valid JSON block
    if text.startswith('{'):
        depth = 0
        for idx, char in enumerate(text):
            if char == '{':
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0:
                    return text[:idx+1]
    return text

def main():
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        api_key = input("🔑 Gemini API key: ").strip()

    client = genai.Client(api_key=api_key)

    # Upload rubric
    print("📄 Uploading rubric...")
    rubric_file = client.files.upload(file=RUBRIC_PDF)
    print(f"   ✅ Rubric uploaded: {rubric_file.name}")

    # Load first student
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    first_row = None
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        name = row[COL_NAME - 1].value
        link_cell = row[COL_LINK - 1]
        if name and link_cell.hyperlink:
            first_row = (str(name).strip(), link_cell.hyperlink.target)
            break

    if not first_row:
        print("❌ No student found with a resume link.")
        return

    student_name, url = first_row
    print(f"\n🎓 Testing with: {student_name}")
    print(f"   URL: {url[:80]}...")

    # Download
    print("   ⬇️  Downloading PDF...")
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    pdf_bytes = resp.content
    print(f"   ✅ Downloaded {len(pdf_bytes):,} bytes")

    # Upload resume to Gemini
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        print("   📄 Uploading resume to Gemini...")
        resume_file = client.files.upload(file=tmp_path)
        print(f"   ✅ Resume uploaded: {resume_file.name}")
    finally:
        os.unlink(tmp_path)

    # Evaluate
    print("   🤖 Asking Gemini...")
    system_instruction = """You are an expert resume reviewer for engineering student placements (branches include electronics, IT, and general engineering — apply judgment broadly across fields).
Resumes are submitted through Superset, a placement platform — so formatting/layout is largely enforced by the platform, not the student. Focus your evaluation on content, not layout. However, the resume must be strictly one page.

You will receive:
1. A rubric PDF explaining the evaluation criteria (formatting, BITS template rules, keywords, action verb guidance).
2. A student's resume PDF.

Your task:
- Carefully evaluate the resume content based on the 100-point rubric below, deducting for any red flags, and determine the score and verdict.
- Note: The current evaluation month and year are provided in the user request. Do not flag dates on the resume as "future dates" unless they are strictly later than that current evaluation date.
- Output a JSON object with exactly these keys:
  {
    "score": <integer 0-100>,
    "verdict": <"green" | "yellow" | "red">,
    "remarks": "<strict remarks following the Remark Writing Rules below>"
  }

RUBRIC (100 pts total):
1. Structure/Field Completeness (10 pts)
   - All relevant Superset fields filled meaningfully — no "N/A," blank sections, or leftover placeholder text where real content should be.
2. Content Quality — Action Verbs & Quantification (30 pts)
   - Bullets start with strong action verbs, not passive phrases like "responsible for / helped / worked on" (10 pts)
   - Achievements are quantified — %, ₹/$, time, scale — not vague claims (10 pts)
   - Implicit STAR logic present: context + action + measurable result, even if not rigidly labeled (10 pts)
3. Projects & Experience Depth (25 pts)
   - Real technical/functional depth, not a surface-level listing of technologies (10 pts)
   - Relevance to target roles (electronics/IT-adjacent or general engineering) (10 pts)
   - Internships/PORs show progression or ownership, not just attendance (5 pts)
4. Skills & Academic Fit (10 pts)
   - Skills are relevant and specific, not generic keyword-stuffing (5 pts)
   - CGPA/coursework represented appropriately (5 pts)
5. Overall Human Read — Gut Check (25 pts)
   Step back and ask: "If I were a recruiter skimming this for 20 seconds, would it land well?"
   - Feels polished, confident, ready to send as-is (12 pts)
   - Tells a coherent story about the candidate, rather than reading like a checklist was filled in mechanically (13 pts)
   *Note: This section is deliberately holistic. If the mechanical score from sections 1-4 doesn't match how the resume actually reads overall, let this section pull the total up or down by a few points to correct for it. A resume that's slightly imperfect on individual bullets but reads strong, clean, and intentional as a whole should score WELL here — do not penalize the same minor flaw twice.

RED FLAGS — deduct 5 pts EACH, stacks, no cap:
   - Unprofessional email address
   - Unexplained timeline gaps
   - Factual inconsistencies (e.g. dates that are in the future relative to the current evaluation date/month)
   - A core section left essentially empty
   - Obvious leftover placeholder/template text
   - Content clearly overcrowded/cramped to force-fit the one-page limit

COLOR MAPPING & VERDICTS (based on final total after deductions):
   - green  : 90-100  (Excellent/Good)
   - yellow : 55-89   (Average/Needs improvement)
   - red    : below 55 (Poor)

REMARK WRITING RULES (strict):
   - Length: 2-3 sentences, under ~50 words total.
   - Every sentence must reference something SPECIFIC from this resume's actual text (a project name, a skill, a missing section) — never generic praise/criticism like "good resume" or "needs improvement."
   - Structure:
     (1) Sentence 1: what's working, if anything is genuinely strong
     (2) Sentence 2: the single biggest gap, naming which rubric category it falls under (e.g. Structure/Field Completeness, Content Quality, Projects & Experience Depth, Skills & Academic Fit, Overall Human Read)
     (3) Sentence 3: one concrete, actionable fix
   - Do NOT restate the score or color inside the remark.
   - If red flags were triggered, name which one(s) explicitly.

Output ONLY the JSON object, no extra text.
"""

    
    from datetime import datetime
    current_date = datetime.now().strftime("%B %Y")
    
    response = client.models.generate_content(
        model="gemini-3.1-flash-lite",
        contents=[
            rubric_file,
            resume_file,
            f"Evaluate the resume of student: {student_name}. Current date context for evaluation: {current_date}. All dates on the resume up to this date (including {current_date}) are in the past or present, NOT the future. Return ONLY a JSON object."
        ],
        config=types.GenerateContentConfig(
            system_instruction=system_instruction,
            response_mime_type="application/json"
        )
    )
    text = response.text.strip()
    print("\n📋 Raw response:")
    print(text)

    cleaned_text = clean_json_text(text)
    result = json.loads(cleaned_text)
    print("\n✅ Parsed result:")
    print(f"   Score  : {result['score']}")
    print(f"   Verdict: {result['verdict']}")
    print(f"   Remarks: {result['remarks']}")
    print("\n🎉 Test passed! Run review_resumes.py for the full batch.")

if __name__ == "__main__":
    main()
